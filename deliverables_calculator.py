"""
Deliverables Calculator Module
Calculates physical delivery obligations for futures and options positions
Works with both pre-trade and post-trade positions
"""

import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List, Optional, Tuple
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import logging

logger = logging.getLogger(__name__)


class DeliverableCalculator:
    """Calculate physical deliverables for futures and options positions"""
    
    def __init__(self, usdinr_rate: float = 88.0):
        self.usdinr_rate = usdinr_rate
        
        # Styles for Excel
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.header_alignment = Alignment(horizontal="center", vertical="center")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        self.group_font = Font(bold=True, size=10)
        self.group_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        
        # Number formats
        self.price_format = '#,##0.00'
        self.deliv_format = '#,##0'
        self.iv_format = '#,##0'
    
    def calculate_deliverables_from_dataframe(self, positions_df: pd.DataFrame, 
                                             prices: Dict[str, float] = None) -> pd.DataFrame:
        """
        Calculate deliverables from a positions DataFrame
        
        Args:
            positions_df: DataFrame with columns [Ticker, Symbol, Security_Type, Strike, Lots, Lot_Size, etc.]
            prices: Optional dictionary of symbol prices
            
        Returns:
            DataFrame with deliverable calculations
        """
        if positions_df.empty:
            return pd.DataFrame()
        
        deliverables_data = []
        
        for idx, row in positions_df.iterrows():
            ticker = row.get('Ticker', '')
            symbol = row.get('Symbol', '')
            security_type = row.get('Security_Type', '')
            strike = float(row.get('Strike', 0))
            lots = float(row.get('Lots', 0))
            lot_size = int(row.get('Lot_Size', 1))
            
            # Get price if available
            spot_price = 0
            if prices and symbol in prices:
                spot_price = prices[symbol]
            elif prices and ticker in prices:
                spot_price = prices[ticker]
            
            # Calculate deliverable based on security type
            if security_type == 'Futures':
                deliverable = lots
            elif security_type == 'Call':
                if spot_price > 0 and spot_price > strike:
                    deliverable = lots
                else:
                    deliverable = 0
            elif security_type == 'Put':
                if spot_price > 0 and spot_price < strike:
                    deliverable = -lots  # Negative for put delivery
                else:
                    deliverable = 0
            else:
                deliverable = 0
            
            # Calculate intrinsic value
            iv = self._calculate_intrinsic_value(security_type, spot_price, strike, lots, lot_size)
            
            deliverables_data.append({
                'Ticker': ticker,
                'Symbol': symbol,
                'Security_Type': security_type,
                'Strike': strike,
                'Lots': lots,
                'Lot_Size': lot_size,
                'Spot_Price': spot_price,
                'Deliverable_Lots': deliverable,
                'Deliverable_Qty': deliverable * lot_size,
                'Intrinsic_Value_INR': iv,
                'Intrinsic_Value_USD': iv / self.usdinr_rate
            })
        
        return pd.DataFrame(deliverables_data)
    
    def _calculate_intrinsic_value(self, security_type: str, spot_price: float, 
                                  strike: float, lots: float, lot_size: int) -> float:
        """Calculate intrinsic value for an option position"""
        if security_type == 'Futures' or spot_price <= 0:
            return 0
        
        if security_type == 'Call':
            if spot_price > strike:
                return lots * lot_size * (spot_price - strike)
        elif security_type == 'Put':
            if spot_price < strike:
                return lots * lot_size * (strike - spot_price)
        
        return 0
    
    def generate_deliverables_report(self, 
                                    starting_positions_df: pd.DataFrame,
                                    final_positions_df: pd.DataFrame,
                                    prices: Dict[str, float],
                                    output_file: str,
                                    report_type: str = "TRADE_PROCESSING") -> str:
        """
        Generate comprehensive deliverables report with multiple sheets
        
        Args:
            starting_positions_df: Pre-trade positions
            final_positions_df: Post-trade positions
            prices: Dictionary of current prices
            output_file: Output Excel filename
            report_type: Type of report for naming convention
            
        Returns:
            Path to generated Excel file
        """
        wb = Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            wb.remove(wb['Sheet'])
        
        # Calculate deliverables for both position sets
        starting_deliv = self.calculate_deliverables_from_dataframe(starting_positions_df, prices)
        final_deliv = self.calculate_deliverables_from_dataframe(final_positions_df, prices)
        
        # 1. Summary Sheet
        ws_summary = wb.create_sheet("Summary")
        self._write_summary_sheet(ws_summary, starting_deliv, final_deliv, report_type)
        
        # 2. Pre-Trade Deliverables
        if not starting_deliv.empty:
            ws_pre = wb.create_sheet("Pre_Trade_Deliverables")
            self._write_deliverables_sheet(ws_pre, starting_deliv, "Pre-Trade Positions")
        
        # 3. Post-Trade Deliverables  
        if not final_deliv.empty:
            ws_post = wb.create_sheet("Post_Trade_Deliverables")
            self._write_deliverables_sheet(ws_post, final_deliv, "Post-Trade Positions")
        
        # 4. Comparison Sheet
        if not starting_deliv.empty and not final_deliv.empty:
            ws_comp = wb.create_sheet("Deliverables_Comparison")
            self._write_comparison_sheet(ws_comp, starting_deliv, final_deliv)
        
        # 5. Pre-Trade IV Sheet
        if not starting_deliv.empty:
            ws_iv_pre = wb.create_sheet("Pre_Trade_IV")
            self._write_iv_sheet(ws_iv_pre, starting_deliv, "Pre-Trade Intrinsic Values")
        
        # 6. Post-Trade IV Sheet
        if not final_deliv.empty:
            ws_iv_post = wb.create_sheet("Post_Trade_IV")
            self._write_iv_sheet(ws_iv_post, final_deliv, "Post-Trade Intrinsic Values")
        
        # Save workbook
        wb.save(output_file)
        logger.info(f"Deliverables report saved: {output_file}")
        return output_file
    
    def _write_summary_sheet(self, ws, starting_deliv: pd.DataFrame, 
                           final_deliv: pd.DataFrame, report_type: str):
        """Write summary sheet with key metrics"""
        ws.cell(row=1, column=1, value=f"{report_type} DELIVERABLES SUMMARY").font = Font(bold=True, size=14)
        ws.cell(row=2, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        ws.cell(row=3, column=1, value=f"USD/INR Rate: {self.usdinr_rate}")
        
        row = 5
        ws.cell(row=row, column=1, value="PRE-TRADE SUMMARY").font = Font(bold=True, size=12)
        row += 1
        
        if not starting_deliv.empty:
            total_pre_deliv = starting_deliv['Deliverable_Lots'].sum()
            total_pre_iv = starting_deliv['Intrinsic_Value_INR'].sum()
            
            ws.cell(row=row, column=1, value="Total Positions:")
            ws.cell(row=row, column=2, value=len(starting_deliv))
            row += 1
            
            ws.cell(row=row, column=1, value="Net Deliverable (Lots):")
            ws.cell(row=row, column=2, value=total_pre_deliv)
            row += 1
            
            ws.cell(row=row, column=1, value="Total Intrinsic Value (INR):")
            ws.cell(row=row, column=2, value=total_pre_iv)
            ws.cell(row=row, column=2).number_format = self.iv_format
            row += 1
            
            ws.cell(row=row, column=1, value="Total Intrinsic Value (USD):")
            ws.cell(row=row, column=2, value=total_pre_iv / self.usdinr_rate)
            ws.cell(row=row, column=2).number_format = self.iv_format
        else:
            ws.cell(row=row, column=1, value="No pre-trade positions")
        
        row += 2
        ws.cell(row=row, column=1, value="POST-TRADE SUMMARY").font = Font(bold=True, size=12)
        row += 1
        
        if not final_deliv.empty:
            total_post_deliv = final_deliv['Deliverable_Lots'].sum()
            total_post_iv = final_deliv['Intrinsic_Value_INR'].sum()
            
            ws.cell(row=row, column=1, value="Total Positions:")
            ws.cell(row=row, column=2, value=len(final_deliv))
            row += 1
            
            ws.cell(row=row, column=1, value="Net Deliverable (Lots):")
            ws.cell(row=row, column=2, value=total_post_deliv)
            row += 1
            
            ws.cell(row=row, column=1, value="Total Intrinsic Value (INR):")
            ws.cell(row=row, column=2, value=total_post_iv)
            ws.cell(row=row, column=2).number_format = self.iv_format
            row += 1
            
            ws.cell(row=row, column=1, value="Total Intrinsic Value (USD):")
            ws.cell(row=row, column=2, value=total_post_iv / self.usdinr_rate)
            ws.cell(row=row, column=2).number_format = self.iv_format
        else:
            ws.cell(row=row, column=1, value="No post-trade positions")
        
        # Changes summary
        if not starting_deliv.empty and not final_deliv.empty:
            row += 2
            ws.cell(row=row, column=1, value="CHANGES").font = Font(bold=True, size=12)
            row += 1
            
            deliv_change = final_deliv['Deliverable_Lots'].sum() - starting_deliv['Deliverable_Lots'].sum()
            iv_change = final_deliv['Intrinsic_Value_INR'].sum() - starting_deliv['Intrinsic_Value_INR'].sum()
            
            ws.cell(row=row, column=1, value="Deliverable Change (Lots):")
            ws.cell(row=row, column=2, value=deliv_change)
            if deliv_change > 0:
                ws.cell(row=row, column=2).font = Font(color="FF0000")  # Red for increase
            elif deliv_change < 0:
                ws.cell(row=row, column=2).font = Font(color="00FF00")  # Green for decrease
            row += 1
            
            ws.cell(row=row, column=1, value="IV Change (INR):")
            ws.cell(row=row, column=2, value=iv_change)
            ws.cell(row=row, column=2).number_format = self.iv_format
        
        # Set column widths
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _write_deliverables_sheet(self, ws, deliv_df: pd.DataFrame, title: str):
        """Write deliverables detail sheet"""
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)
        
        # Write headers
        headers = ['Ticker', 'Symbol', 'Type', 'Strike', 'Lots', 'Lot Size',
                  'Spot Price', 'Deliverable (Lots)', 'Deliverable (Qty)']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Write data
        row = 4
        for idx, data_row in deliv_df.iterrows():
            ws.cell(row=row, column=1, value=data_row['Ticker'])
            ws.cell(row=row, column=2, value=data_row['Symbol'])
            ws.cell(row=row, column=3, value=data_row['Security_Type'])
            ws.cell(row=row, column=4, value=data_row['Strike']).number_format = self.price_format
            ws.cell(row=row, column=5, value=data_row['Lots'])
            ws.cell(row=row, column=6, value=data_row['Lot_Size'])
            ws.cell(row=row, column=7, value=data_row['Spot_Price']).number_format = self.price_format
            ws.cell(row=row, column=8, value=data_row['Deliverable_Lots'])
            ws.cell(row=row, column=9, value=data_row['Deliverable_Qty'])
            
            # Apply borders
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = self.border
            
            # Color code deliverables
            if data_row['Deliverable_Lots'] > 0:
                ws.cell(row=row, column=8).fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
            elif data_row['Deliverable_Lots'] < 0:
                ws.cell(row=row, column=8).fill = PatternFill(start_color="E5FFE5", end_color="E5FFE5", fill_type="solid")
            
            row += 1
        
        # Add totals row
        ws.cell(row=row+1, column=7, value="TOTAL").font = self.header_font
        ws.cell(row=row+1, column=8, value=deliv_df['Deliverable_Lots'].sum()).font = self.header_font
        ws.cell(row=row+1, column=9, value=deliv_df['Deliverable_Qty'].sum()).font = self.header_font
        
        # Set column widths
        col_widths = {'A': 30, 'B': 15, 'C': 10, 'D': 10, 'E': 10, 
                     'F': 10, 'G': 12, 'H': 18, 'I': 18}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
    
    def _write_iv_sheet(self, ws, deliv_df: pd.DataFrame, title: str):
        """Write intrinsic value detail sheet"""
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=12)
        
        # Write headers
        headers = ['Ticker', 'Symbol', 'Type', 'Strike', 'Lots', 'Lot Size',
                  'Spot Price', 'IV (INR)', 'IV (USD)']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Filter only options with IV > 0
        iv_df = deliv_df[deliv_df['Intrinsic_Value_INR'] > 0].copy()
        
        # Write data
        row = 4
        for idx, data_row in iv_df.iterrows():
            ws.cell(row=row, column=1, value=data_row['Ticker'])
            ws.cell(row=row, column=2, value=data_row['Symbol'])
            ws.cell(row=row, column=3, value=data_row['Security_Type'])
            ws.cell(row=row, column=4, value=data_row['Strike']).number_format = self.price_format
            ws.cell(row=row, column=5, value=data_row['Lots'])
            ws.cell(row=row, column=6, value=data_row['Lot_Size'])
            ws.cell(row=row, column=7, value=data_row['Spot_Price']).number_format = self.price_format
            ws.cell(row=row, column=8, value=data_row['Intrinsic_Value_INR']).number_format = self.iv_format
            ws.cell(row=row, column=9, value=data_row['Intrinsic_Value_USD']).number_format = self.iv_format
            
            # Apply borders
            for col in range(1, 10):
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        # Add totals row
        if not iv_df.empty:
            ws.cell(row=row+1, column=7, value="TOTAL").font = self.header_font
            ws.cell(row=row+1, column=8, value=iv_df['Intrinsic_Value_INR'].sum()).font = self.header_font
            ws.cell(row=row+1, column=8).number_format = self.iv_format
            ws.cell(row=row+1, column=9, value=iv_df['Intrinsic_Value_USD'].sum()).font = self.header_font
            ws.cell(row=row+1, column=9).number_format = self.iv_format
        
        # Set column widths
        col_widths = {'A': 30, 'B': 15, 'C': 10, 'D': 10, 'E': 10,
                     'F': 10, 'G': 12, 'H': 18, 'I': 18}
        for col, width in col_widths.items():
            ws.column_dimensions[col].width = width
    
    def _write_comparison_sheet(self, ws, starting_deliv: pd.DataFrame, final_deliv: pd.DataFrame):
        """Write comparison sheet showing changes"""
        ws.cell(row=1, column=1, value="DELIVERABLES COMPARISON").font = Font(bold=True, size=12)
        
        # Merge dataframes on Ticker
        merged = pd.merge(
            starting_deliv[['Ticker', 'Symbol', 'Security_Type', 'Strike', 'Deliverable_Lots', 'Intrinsic_Value_INR']],
            final_deliv[['Ticker', 'Deliverable_Lots', 'Intrinsic_Value_INR']],
            on='Ticker',
            how='outer',
            suffixes=('_Pre', '_Post')
        )
        
        # Fill NaN values
        merged = merged.fillna(0)
        
        # Calculate changes
        merged['Deliv_Change'] = merged['Deliverable_Lots_Post'] - merged['Deliverable_Lots_Pre']
        merged['IV_Change'] = merged['Intrinsic_Value_INR_Post'] - merged['Intrinsic_Value_INR_Pre']
        
        # Write headers
        headers = ['Ticker', 'Symbol', 'Type', 'Strike', 
                  'Pre-Trade Deliv', 'Post-Trade Deliv', 'Deliv Change',
                  'Pre-Trade IV', 'Post-Trade IV', 'IV Change']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.alignment = self.header_alignment
            cell.border = self.border
        
        # Write data
        row = 4
        for idx, data_row in merged.iterrows():
            ws.cell(row=row, column=1, value=data_row['Ticker'])
            ws.cell(row=row, column=2, value=data_row.get('Symbol', ''))
            ws.cell(row=row, column=3, value=data_row.get('Security_Type', ''))
            ws.cell(row=row, column=4, value=data_row.get('Strike', 0))
            ws.cell(row=row, column=5, value=data_row['Deliverable_Lots_Pre'])
            ws.cell(row=row, column=6, value=data_row['Deliverable_Lots_Post'])
            ws.cell(row=row, column=7, value=data_row['Deliv_Change'])
            ws.cell(row=row, column=8, value=data_row['Intrinsic_Value_INR_Pre']).number_format = self.iv_format
            ws.cell(row=row, column=9, value=data_row['Intrinsic_Value_INR_Post']).number_format = self.iv_format
            ws.cell(row=row, column=10, value=data_row['IV_Change']).number_format = self.iv_format
            
            # Color code changes
            if data_row['Deliv_Change'] != 0:
                if data_row['Deliv_Change'] > 0:
                    ws.cell(row=row, column=7).fill = PatternFill(start_color="FFE5E5", end_color="FFE5E5", fill_type="solid")
                else:
                    ws.cell(row=row, column=7).fill = PatternFill(start_color="E5FFE5", end_color="E5FFE5", fill_type="solid")
            
            # Apply borders
            for col in range(1, 11):
                ws.cell(row=row, column=col).border = self.border
            
            row += 1
        
        # Set column widths
        for col in ['A', 'B']:
            ws.column_dimensions[col].width = 25
        for col in ['C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']:
            ws.column_dimensions[col].width = 15
