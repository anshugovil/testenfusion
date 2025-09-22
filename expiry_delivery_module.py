"""
Expiry Delivery Generator Module
Generates physical delivery trades and cash settlements per expiry date
Integrates with existing Trade Processing Pipeline
"""

import pandas as pd
import numpy as np
from datetime import datetime
from typing import Dict, List, Tuple, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
from pathlib import Path

logger = logging.getLogger(__name__)


class ExpiryDeliveryGenerator:
    """Generate physical delivery trades for expiring positions"""
    
    def __init__(self, usdinr_rate: float = 88.0):
        self.usdinr_rate = usdinr_rate
        
        # Excel styles
        self.header_font = Font(bold=True, size=11)
        self.header_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
        self.highlight_fill = PatternFill(start_color="ADD8E6", end_color="ADD8E6", fill_type="solid")
        self.total_fill = PatternFill(start_color="FFD700", end_color="FFD700", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    def process_positions_by_expiry(self, 
                                   positions_df: pd.DataFrame, 
                                   prices: Dict[str, float],
                                   position_type: str = "Post-Trade") -> Dict[datetime, Dict]:
        """
        Process positions grouped by expiry date
        
        Args:
            positions_df: DataFrame with columns [Ticker, Symbol, Security_Type, Expiry, Strike, Lots, Lot_Size, etc.]
            prices: Dictionary of symbol -> price from Yahoo
            position_type: "Pre-Trade" or "Post-Trade"
            
        Returns:
            Dictionary with expiry date as key and processing results as value
        """
        if positions_df.empty:
            return {}
        
        # Ensure Expiry is datetime
        positions_df = positions_df.copy()
        positions_df['Expiry'] = pd.to_datetime(positions_df['Expiry'])
        
        # Group by expiry date
        expiry_groups = positions_df.groupby(positions_df['Expiry'].dt.date)
        
        results = {}
        
        for expiry_date, group_df in expiry_groups:
            logger.info(f"Processing {len(group_df)} positions for expiry {expiry_date}")
            
            # Process this expiry group
            derivatives, cash_trades, cash_summary, errors = self._process_expiry_group(
                group_df, prices, expiry_date
            )
            
            results[expiry_date] = {
                'position_type': position_type,
                'expiry_date': expiry_date,
                'derivatives': derivatives,
                'cash_trades': cash_trades,
                'cash_summary': cash_summary,
                'errors': errors,
                'position_count': len(group_df)
            }
        
        return results
    
    def _process_expiry_group(self, 
                            group_df: pd.DataFrame, 
                            prices: Dict[str, float],
                            expiry_date) -> Tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, pd.DataFrame]:
        """Process a single expiry group"""
        derivatives_list = []
        cash_trades_list = []
        errors_list = []
        
        for idx, row in group_df.iterrows():
            try:
                # Get price for this position
                last_price = self._get_price_for_position(row, prices)
                
                if last_price is None:
                    errors_list.append({
                        'Symbol': row['Symbol'],
                        'Ticker': row['Ticker'],
                        'Reason': 'No price available',
                        'Expiry': expiry_date
                    })
                    continue
                
                # Process based on security type
                if row['Security_Type'] == 'Futures':
                    deriv, cash = self._process_futures(row, last_price)
                elif row['Security_Type'] in ['Call', 'Put']:
                    deriv, cash = self._process_option(row, last_price)
                else:
                    errors_list.append({
                        'Symbol': row['Symbol'],
                        'Ticker': row['Ticker'],
                        'Reason': f"Unknown security type: {row['Security_Type']}",
                        'Expiry': expiry_date
                    })
                    continue
                
                if deriv:
                    derivatives_list.append(deriv)
                if cash:
                    cash_trades_list.append(cash)
                    
            except Exception as e:
                errors_list.append({
                    'Symbol': row.get('Symbol', 'N/A'),
                    'Ticker': row.get('Ticker', 'N/A'),
                    'Reason': str(e),
                    'Expiry': expiry_date
                })
        
        # Create DataFrames
        derivatives_df = pd.DataFrame(derivatives_list) if derivatives_list else pd.DataFrame()
        cash_df = pd.DataFrame(cash_trades_list) if cash_trades_list else pd.DataFrame()
        errors_df = pd.DataFrame(errors_list) if errors_list else pd.DataFrame()
        
        # Generate cash summary
        cash_summary_df = self._generate_cash_summary(cash_df) if not cash_df.empty else pd.DataFrame()
        
        return derivatives_df, cash_df, cash_summary_df, errors_df
    
    def _get_price_for_position(self, row: pd.Series, prices: Dict[str, float]) -> Optional[float]:
        """Get price for a position from Yahoo prices or stored prices"""
        # First try Yahoo_Price column if it exists
        if 'Yahoo_Price' in row and pd.notna(row['Yahoo_Price']) and row['Yahoo_Price'] != 'N/A':
            try:
                return float(row['Yahoo_Price'])
            except:
                pass
        
        # Then try the prices dictionary with various keys
        for key in ['Symbol', 'Underlying']:
            if key in row:
                symbol = row[key]
                if symbol in prices:
                    return prices[symbol]
                
                # Try removing suffixes
                base_symbol = str(symbol).split(' ')[0] if pd.notna(symbol) else None
                if base_symbol and base_symbol in prices:
                    return prices[base_symbol]
        
        return None
    
    def _is_index_product(self, ticker: str) -> bool:
        """Check if the ticker is an index product"""
        if pd.isna(ticker):
            return False
        ticker_upper = str(ticker).upper()
        return 'INDEX' in ticker_upper or any(idx in ticker_upper for idx in ['NIFTY', 'NZ', 'AF1', 'NSEBANK'])
    
    def _process_futures(self, row: pd.Series, last_price: float) -> Tuple[Dict, Optional[Dict]]:
        """Process futures position at expiry"""
        position = float(row['Lots'])
        lot_size = float(row['Lot_Size'])
        ticker = row['Ticker']
        symbol = row['Symbol']
        
        is_index = self._is_index_product(ticker)
        
        # Determine underlying (extract from Bloomberg ticker if needed)
        if 'Underlying' in row and pd.notna(row['Underlying']):
            underlying = row['Underlying']
        else:
            # Extract from ticker
            if ' IS Equity' in ticker:
                underlying = ticker.split('=')[0] + ' IS Equity'
            elif ' Index' in ticker:
                underlying = ticker.split(' ')[0] + ' Index'
            else:
                underlying = symbol
        
        # Derivatives entry (closing the futures)
        derivative = {
            'Underlying': underlying,
            'Symbol': ticker,
            'Expiry': row['Expiry'],
            'Buy/Sell': 'Sell' if position > 0 else 'Buy',
            'Strategy': 'FULO' if position > 0 else 'FUSH',
            'Position': abs(position),
            'Price': last_price,
            'Type': 'Futures',
            'Strike': '',
            'Lot Size': lot_size,
            'tradenotes': ''
        }
        
        # Cash entry - only for stock futures, not index futures
        cash = None
        if not is_index:
            cash_quantity = abs(position) * lot_size
            
            # Tax Calculations for futures
            stt = cash_quantity * last_price * 0.001  # 0.1% STT
            stamp_duty = cash_quantity * last_price * 0.00002  # 0.002% stamp duty
            taxes = stt + stamp_duty
            
            cash = {
                'Underlying': underlying,
                'Symbol': underlying,
                'Expiry': '',
                'Buy/Sell': 'Buy' if position > 0 else 'Sell',
                'Strategy': 'EQLO2',
                'Position': cash_quantity,
                'Price': last_price,
                'Type': 'CASH',
                'Strike': '',
                'Lot Size': '',
                'tradenotes': '',  # Blank for futures
                'STT': round(stt, 2),
                'Stamp Duty': round(stamp_duty, 2),
                'Taxes': round(taxes, 2)
            }
        
        return derivative, cash
    
    def _process_option(self, row: pd.Series, last_price: float) -> Tuple[Dict, Optional[Dict]]:
        """Process option position at expiry"""
        position = float(row['Lots'])
        lot_size = float(row['Lot_Size'])
        strike = float(row['Strike']) if pd.notna(row['Strike']) else 0
        option_type = row['Security_Type']  # 'Call' or 'Put'
        ticker = row['Ticker']
        symbol = row['Symbol']
        
        is_index = self._is_index_product(ticker)
        
        # Determine underlying
        if 'Underlying' in row and pd.notna(row['Underlying']):
            underlying = row['Underlying']
        else:
            # Extract from ticker
            if ' IS ' in ticker:
                underlying = ticker.split(' IS ')[0] + ' IS Equity'
            elif ' Index' in ticker:
                parts = ticker.split(' ')
                underlying = parts[0] + ' Index'
            else:
                underlying = symbol
        
        # Determine if ITM
        is_itm = self._is_option_itm(option_type, strike, last_price)
        
        # Derivatives entry
        if option_type == 'Call':
            deriv_buy_sell = 'Sell' if position > 0 else 'Buy'
            deriv_strategy = 'FULO' if position > 0 else 'FUSH'
        else:  # Put
            deriv_buy_sell = 'Sell' if position > 0 else 'Buy'
            deriv_strategy = 'FUSH' if position > 0 else 'FULO'
        
        # Determine price for derivatives
        if is_index and is_itm:
            # Index options cash settle to intrinsic value
            if option_type == 'Call':
                deriv_price = max(0, last_price - strike)
            else:
                deriv_price = max(0, strike - last_price)
        else:
            deriv_price = 0
        
        # Determine tradenotes for derivatives
        tradenotes = ''
        if is_itm and not is_index:
            if deriv_buy_sell == 'Buy':
                tradenotes = 'A'  # Assignment (we were short, now buying back)
            else:
                tradenotes = 'E'  # Exercise (we were long, now selling)
        
        derivative = {
            'Underlying': underlying,
            'Symbol': ticker,
            'Expiry': row['Expiry'],
            'Buy/Sell': deriv_buy_sell,
            'Strategy': deriv_strategy,
            'Position': abs(position),
            'Price': deriv_price,
            'Type': option_type,
            'Strike': strike,
            'Lot Size': lot_size,
            'tradenotes': tradenotes
        }
        
        # Cash entry - only for ITM single stock options
        cash = None
        if is_itm and not is_index:
            cash_quantity = abs(position) * lot_size
            
            if option_type == 'Call':
                # Call: Long calls buy stock at strike, Short calls sell stock at strike
                cash_buy_sell = 'Buy' if position > 0 else 'Sell'
                cash_price = strike
                intrinsic_value = last_price - strike
            else:  # Put
                # Put: Long puts sell stock at strike, Short puts buy stock at strike
                cash_buy_sell = 'Sell' if position > 0 else 'Buy'
                cash_price = strike
                intrinsic_value = strike - last_price
            
            # Tax Calculations - only long options pay taxes
            if position > 0:  # Long options being exercised
                stt = cash_quantity * max(0, intrinsic_value) * 0.00125  # 0.125% of intrinsic
                stamp_duty = cash_quantity * strike * 0.00003  # 0.003% of strike
            else:  # Short options being assigned - no taxes
                stt = 0
                stamp_duty = 0
            
            taxes = stt + stamp_duty
            
            # Tradenotes for cash: opposite of derivatives
            cash_tradenotes = 'E' if position > 0 else 'A'
            
            cash = {
                'Underlying': underlying,
                'Symbol': underlying,
                'Expiry': '',
                'Buy/Sell': cash_buy_sell,
                'Strategy': 'EQLO2',
                'Position': cash_quantity,
                'Price': cash_price,
                'Type': 'CASH',
                'Strike': '',
                'Lot Size': '',
                'tradenotes': cash_tradenotes,
                'STT': round(stt, 2),
                'Stamp Duty': round(stamp_duty, 2),
                'Taxes': round(taxes, 2)
            }
        
        return derivative, cash
    
    def _is_option_itm(self, option_type: str, strike: float, spot_price: float) -> bool:
        """Determine if option is in-the-money"""
        if option_type == 'Call':
            return spot_price > strike
        elif option_type == 'Put':
            return spot_price < strike
        return False
    
    def _generate_cash_summary(self, cash_df: pd.DataFrame) -> pd.DataFrame:
        """Generate cash summary with net deliverables per underlying"""
        if cash_df.empty:
            return pd.DataFrame()
        
        summary_rows = []
        
        # Variables for grand totals
        grand_total_consideration = 0
        grand_total_stt = 0
        grand_total_stamp = 0
        grand_total_taxes = 0
        
        # Group by underlying
        for underlying in cash_df['Underlying'].unique():
            underlying_trades = cash_df[cash_df['Underlying'] == underlying].copy()
            
            # Add individual trade rows
            for idx, trade in underlying_trades.iterrows():
                quantity = trade['Position']
                price = trade['Price']
                consideration = quantity * price if trade['Buy/Sell'] == 'Buy' else -quantity * price
                
                summary_rows.append({
                    'Underlying': underlying,
                    'Type': 'Trade',
                    'Buy/Sell': trade['Buy/Sell'],
                    'Quantity': quantity,
                    'Price': price,
                    'Consideration': round(consideration, 2),
                    'STT': trade.get('STT', 0),
                    'Stamp Duty': trade.get('Stamp Duty', 0),
                    'Taxes': trade.get('Taxes', 0),
                    'TradeNotes': trade.get('tradenotes', '')
                })
            
            # Calculate net deliverable for this underlying
            buy_trades = underlying_trades[underlying_trades['Buy/Sell'] == 'Buy']
            sell_trades = underlying_trades[underlying_trades['Buy/Sell'] == 'Sell']
            
            buy_qty = buy_trades['Position'].sum() if not buy_trades.empty else 0
            sell_qty = sell_trades['Position'].sum() if not sell_trades.empty else 0
            net_qty = buy_qty - sell_qty
            
            buy_consideration = sum(row['Position'] * row['Price'] for _, row in buy_trades.iterrows()) if not buy_trades.empty else 0
            sell_consideration = sum(row['Position'] * row['Price'] for _, row in sell_trades.iterrows()) if not sell_trades.empty else 0
            net_consideration = buy_consideration - sell_consideration
            
            total_stt = underlying_trades['STT'].sum()
            total_stamp = underlying_trades['Stamp Duty'].sum()
            total_taxes = underlying_trades['Taxes'].sum()
            
            # Add to grand totals
            grand_total_consideration += net_consideration
            grand_total_stt += total_stt
            grand_total_stamp += total_stamp
            grand_total_taxes += total_taxes
            
            # Add NET DELIVERABLE row
            summary_rows.append({
                'Underlying': underlying,
                'Type': 'NET DELIVERABLE',
                'Buy/Sell': 'NET',
                'Quantity': net_qty,
                'Price': '',
                'Consideration': round(net_consideration, 2),
                'STT': round(total_stt, 2),
                'Stamp Duty': round(total_stamp, 2),
                'Taxes': round(total_taxes, 2),
                'TradeNotes': ''
            })
            
            # Add blank separator row
            if underlying != cash_df['Underlying'].unique()[-1]:
                summary_rows.append({col: '' for col in 
                    ['Underlying', 'Type', 'Buy/Sell', 'Quantity', 'Price', 
                     'Consideration', 'STT', 'Stamp Duty', 'Taxes', 'TradeNotes']})
        
        # Add separator before grand total
        summary_rows.append({col: '---' for col in 
            ['Underlying', 'Type', 'Buy/Sell', 'Quantity', 'Price', 
             'Consideration', 'STT', 'Stamp Duty', 'Taxes', 'TradeNotes']})
        
        # Add GRAND TOTAL row
        summary_rows.append({
            'Underlying': 'GRAND TOTAL',
            'Type': 'ALL POSITIONS',
            'Buy/Sell': '',
            'Quantity': '',
            'Price': '',
            'Consideration': round(grand_total_consideration, 2),
            'STT': round(grand_total_stt, 2),
            'Stamp Duty': round(grand_total_stamp, 2),
            'Taxes': round(grand_total_taxes, 2),
            'TradeNotes': ''
        })
        
        return pd.DataFrame(summary_rows)
    
    def generate_expiry_reports(self,
                              pre_trade_results: Dict,
                              post_trade_results: Dict,
                              output_dir: str) -> Dict[str, str]:
        """
        Generate comprehensive Excel reports for each expiry
        
        Returns:
            Dictionary of expiry_date -> file_path
        """
        output_files = {}
        output_path = Path(output_dir)
        output_path.mkdir(parents=True, exist_ok=True)
        
        # Get all unique expiry dates
        all_expiries = set(list(pre_trade_results.keys()) + list(post_trade_results.keys()))
        
        for expiry_date in sorted(all_expiries):
            # Create workbook for this expiry
            wb = Workbook()
            
            # Remove default sheet
            if 'Sheet' in wb.sheetnames:
                wb.remove(wb['Sheet'])
            
            # Get data for this expiry
            pre_data = pre_trade_results.get(expiry_date, {})
            post_data = post_trade_results.get(expiry_date, {})
            
            # Add sheets based on available data
            sheet_added = False
            
            # Pre-Trade Derivatives
            if pre_data and not pre_data.get('derivatives', pd.DataFrame()).empty:
                ws = wb.create_sheet("PreTrade_Derivatives")
                self._write_dataframe_to_sheet(ws, pre_data['derivatives'], "Pre-Trade Derivatives")
                sheet_added = True
            
            # Pre-Trade Cash
            if pre_data and not pre_data.get('cash_trades', pd.DataFrame()).empty:
                ws = wb.create_sheet("PreTrade_Cash")
                self._write_dataframe_to_sheet(ws, pre_data['cash_trades'], "Pre-Trade Cash Trades")
                sheet_added = True
            
            # Pre-Trade Cash Summary
            if pre_data and not pre_data.get('cash_summary', pd.DataFrame()).empty:
                ws = wb.create_sheet("PreTrade_Summary")
                self._write_cash_summary_sheet(ws, pre_data['cash_summary'], "Pre-Trade Cash Summary")
                sheet_added = True
            
            # Post-Trade Derivatives
            if post_data and not post_data.get('derivatives', pd.DataFrame()).empty:
                ws = wb.create_sheet("PostTrade_Derivatives")
                self._write_dataframe_to_sheet(ws, post_data['derivatives'], "Post-Trade Derivatives")
                sheet_added = True
            
            # Post-Trade Cash
            if post_data and not post_data.get('cash_trades', pd.DataFrame()).empty:
                ws = wb.create_sheet("PostTrade_Cash")
                self._write_dataframe_to_sheet(ws, post_data['cash_trades'], "Post-Trade Cash Trades")
                sheet_added = True
            
            # Post-Trade Cash Summary
            if post_data and not post_data.get('cash_summary', pd.DataFrame()).empty:
                ws = wb.create_sheet("PostTrade_Summary")
                self._write_cash_summary_sheet(ws, post_data['cash_summary'], "Post-Trade Cash Summary")
                sheet_added = True
            
            # Comparison sheet
            if pre_data and post_data:
                ws = wb.create_sheet("Comparison")
                self._write_comparison_sheet(ws, pre_data, post_data)
                sheet_added = True
            
            # Errors sheet
            errors_list = []
            if pre_data and not pre_data.get('errors', pd.DataFrame()).empty:
                errors_df = pre_data['errors'].copy()
                errors_df['Stage'] = 'Pre-Trade'
                errors_list.append(errors_df)
            if post_data and not post_data.get('errors', pd.DataFrame()).empty:
                errors_df = post_data['errors'].copy()
                errors_df['Stage'] = 'Post-Trade'
                errors_list.append(errors_df)
            
            if errors_list:
                combined_errors = pd.concat(errors_list, ignore_index=True)
                ws = wb.create_sheet("Errors")
                self._write_dataframe_to_sheet(ws, combined_errors, "Processing Errors")
                sheet_added = True
            
            # Save workbook if sheets were added
            if sheet_added:
                file_name = f"EXPIRY_DELIVERY_{expiry_date.strftime('%Y%m%d')}.xlsx"
                file_path = output_path / file_name
                wb.save(file_path)
                output_files[expiry_date] = str(file_path)
                logger.info(f"Generated expiry report: {file_path}")
            else:
                logger.warning(f"No data for expiry {expiry_date}, skipping file generation")
        
        return output_files
    
    def _write_dataframe_to_sheet(self, ws, df: pd.DataFrame, title: str):
        """Write DataFrame to worksheet with formatting"""
        # Add title
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
        
        # Add headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Add data
        for row_idx, row in enumerate(df.itertuples(index=False), 4):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def _write_cash_summary_sheet(self, ws, df: pd.DataFrame, title: str):
        """Write cash summary with special formatting"""
        # Add title
        ws.cell(row=1, column=1, value=title).font = Font(bold=True, size=14)
        
        # Add headers
        for col_idx, col_name in enumerate(df.columns, 1):
            cell = ws.cell(row=3, column=col_idx, value=col_name)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
        
        # Add data with special formatting for NET and GRAND TOTAL rows
        for row_idx, row in enumerate(df.itertuples(index=False), 4):
            is_net = 'NET DELIVERABLE' in str(row[1]) if len(row) > 1 else False
            is_total = 'GRAND TOTAL' in str(row[0]) if len(row) > 0 else False
            
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
                
                if is_net:
                    cell.fill = self.highlight_fill
                    cell.font = Font(bold=True)
                elif is_total:
                    cell.fill = self.total_fill
                    cell.font = Font(bold=True, size=12)
        
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column[0].column_letter].width = adjusted_width
    
    def _write_comparison_sheet(self, ws, pre_data: Dict, post_data: Dict):
        """Write comparison between pre and post trade deliveries"""
        ws.cell(row=1, column=1, value="PRE vs POST TRADE COMPARISON").font = Font(bold=True, size=14)
        
        # Summary metrics
        row = 3
        ws.cell(row=row, column=1, value="Metric").font = self.header_font
        ws.cell(row=row, column=2, value="Pre-Trade").font = self.header_font
        ws.cell(row=row, column=3, value="Post-Trade").font = self.header_font
        ws.cell(row=row, column=4, value="Change").font = self.header_font
        
        # Apply header formatting
        for col in range(1, 5):
            cell = ws.cell(row=row, column=col)
            cell.fill = self.header_fill
            cell.border = self.border
        
        row += 1
        
        # Position counts
        pre_positions = len(pre_data.get('derivatives', pd.DataFrame()))
        post_positions = len(post_data.get('derivatives', pd.DataFrame()))
        
        ws.cell(row=row, column=1, value="Derivative Positions")
        ws.cell(row=row, column=2, value=pre_positions)
        ws.cell(row=row, column=3, value=post_positions)
        ws.cell(row=row, column=4, value=post_positions - pre_positions)
        
        row += 1
        
        # Cash trades
        pre_cash = len(pre_data.get('cash_trades', pd.DataFrame()))
        post_cash = len(post_data.get('cash_trades', pd.DataFrame()))
        
        ws.cell(row=row, column=1, value="Cash Trades")
        ws.cell(row=row, column=2, value=pre_cash)
        ws.cell(row=row, column=3, value=post_cash)
        ws.cell(row=row, column=4, value=post_cash - pre_cash)
        
        # Apply borders
        for r in range(3, row + 1):
            for c in range(1, 5):
                ws.cell(row=r, column=c).border = self.border
        
        # Auto-adjust column widths
        for col in ['A', 'B', 'C', 'D']:
            ws.column_dimensions[col].width = 20
